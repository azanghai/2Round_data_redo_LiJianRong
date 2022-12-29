import time
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from DataProcessPack import get_sentiment

dest_filename = r'C:\Users\Hmx\PycharmProjects\2Round_data_redo_LiJianRong\raw_data\douyin_raw_data\douyin_data\douyin_7073782200499571999_2022_11_12_21_56_51.xlsx'
wb = load_workbook(filename=dest_filename)
for st1 in wb:
    print('共{}条数据等待分析.....'.format(st1.max_row))
    print('-' * 20)
    for i in range(1, st1.max_row + 1):
        TextContent = st1.cell(i, 2).value
        try:
            ReturnData = get_sentiment(TextContent)
        except:
            pass
        st1.cell(i, 6).value = ReturnData.get('desc')
        st1.cell(i, 7).value = ReturnData.get('data').get('score')
        st1.cell(i, 8).value = ReturnData.get('data').get('sentiment')
        st1.cell(i, 9).value = ReturnData.get('sid')
        print('第{}条数据分析完成，剩余{}条数据'.format(i - 1, st1.max_row - i))
        print('-' * 20)
        # time.sleep(0.5)
        wb.save(filename=dest_filename)

    wb.save(filename=dest_filename)
    print('所有数据分析完成！')