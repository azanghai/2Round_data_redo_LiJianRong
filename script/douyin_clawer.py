import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl

def get_douyin_cookie():
    cookie_driver = webdriver.Chrome()
    cookie_driver.get(url='https://www.douyin.com/video/7094445702536760590')
    time.sleep(5)
    cookie_driver.execute_script("window.scrollTo(0,document.body.scrollHeight)")
    input('press enter after login')
    cookies = cookie_driver.get_cookies()
    file_name = str(time.strftime("%Y_%m_%d_%H_%M_%S", time.localtime()))
    with open(f'../raw_data/douyin_raw_data/douyin_cookie/{file_name}.txt', 'w') as f:
        f.write(str(cookies))
    print(cookies)
    print(type(cookies))
    # print(cookies[0])
    # print(type(cookies[0]))
    # print(str(cookies))
    # print(str(cookies[0]))
    cookie_driver.close()
    return cookies

def douyin_comment_clawer(douyinid,
                          douyincookie,
                          likenumber,
                          like_class_value,
                          time_class_name,
                          name_class_name,
                          comment_area_class_name,
                          block_class_name,
                          reply_comment_class_name,
                          comment_first,
                          comment_second,
                          at_someone_class):
    driver = webdriver.Chrome()
    Url = f'https://www.douyin.com/video/{douyinid}'
    driver.get(url=Url)
    for singledict in douyincookie:
        driver.add_cookie(singledict)
    print('Cookie add successfully, refresh now')
    driver.refresh()
    input('press enter to continue')
    time.sleep(3)
    driver.execute_script("window.scrollTo(0,document.body.scrollHeight)")
    while True:
        try:
            if int(driver.find_elements(by=By.CLASS_NAME,value=like_class_value)[-1].text) < int(likenumber):
                break
            else:
                iframe = driver.find_elements(by=By.CLASS_NAME, value=like_class_value)[-1]
                ActionChains(driver) \
                    .scroll_to_element(iframe) \
                    .perform()
                # sleep 5 second to avoid being blocked
                time.sleep(5)
        except:
            break
    print('finding complete, begin to save data')

    redo_count = 0
    while True:
        res = driver.find_elements(by=By.CLASS_NAME,value=comment_area_class_name)
        print(res)
        sep_res = res[0].find_elements(by=By.CLASS_NAME, value=block_class_name)

        try:
            reply_comment = res[0].find_elements(by=By.CLASS_NAME, value=reply_comment_class_name)
            reply_list = []
            for single_reply_list in reply_comment:
                delete = single_reply_list.find_element(by=By.CLASS_NAME, value=block_class_name)
                sep_res.remove(delete)
        except:
            pass

        if len(sep_res)>200:
            break
        else:
            iframe = driver.find_elements(by=By.CLASS_NAME, value=like_class_value)[-1]
            ActionChains(driver) \
                .scroll_to_element(iframe) \
                .perform()
            # sleep 5 second to avoid being blocked
            time.sleep(5)
            redo_count = redo_count+1
        if redo_count>20:
            break

    wb = openpyxl.Workbook()
    destfile = f'../raw_data/douyin_raw_data/douyin_data/douyin_{douyinid}_{str(time.strftime("%Y_%m_%d_%H_%M_%S", time.localtime()))}.xlsx'
    ws1 = wb.active
    row_control = 0
    for i in sep_res:
        row_control = row_control + 1
        # print(row_control)
        name = i.find_element(by=By.CLASS_NAME, value=name_class_name)
        ws1.cell(row_control, 1).value = name.text
        comment = i.find_element(by=By.CLASS_NAME, value=comment_first).find_element(by=By.CLASS_NAME,
                                                                                     value=comment_second)
        ws1.cell(row_control, 2).value = comment.text
        try:
            at_someone = i.find_element(by=By.CLASS_NAME, value=at_someone_class)
        except:
            at_someone = 'Null'
        try:
            ws1.cell(row_control,3).value = at_someone.text
        except:
            ws1.cell(row_control,3).value = at_someone
        like = i.find_element(by=By.CLASS_NAME, value=like_class_value)
        ws1.cell(row_control,4).value = like.text
        try:
            comment_time = i.find_element(by=By.CLASS_NAME, value=time_class_name)
            ws1.cell(row_control, 5).value = comment_time.text
        except:
            continue
        print(f'username:{name.text}\ncomment:{comment.text}')
        try:
            print(f'at_someone:{at_someone.text}')
        except:
            print('at_someone:')
        print(f'like number:{like.text}\ncomment time{comment_time.text}\n-----------------')

    wb.save(filename=destfile)
    driver.close()

if __name__ == '__main__':
    # get_douyin_cookie()
    douyincookie = [{'domain': '.douyin.com', 'expiry': 1668784707, 'httpOnly': False, 'name': 'msToken', 'path': '/', 'sameSite': 'None', 'secure': True, 'value': 'uV3Y6VNdDFFCkHMndxqw0FLCdtEgUMvUZmBZSa0e_pH_mFMLWjV7RArz2tuj72uHvT3CsxGj0VmB0zsXKQZoyWtN86zTBC3GrUTan_ysw5hn35eb5i7i9FjxHyt-eIE='}, {'domain': 'www.douyin.com', 'httpOnly': False, 'name': 'csrf_session_id', 'path': '/', 'sameSite': 'None', 'secure': True, 'value': '7c3c43e52ace7ebe6586443fe28d7fe9'}, {'domain': '.douyin.com', 'expiry': 1673363904, 'httpOnly': True, 'name': 'ssid_ucp_sso_v1', 'path': '/', 'sameSite': 'None', 'secure': True, 'value': '1.0.0-KDJhZGY1NzFjNWRhMzkxY2QxNjUyMDdhMDFjZmM2OGExODkwYmQ1ZGMKHQjjhsukhgIQwM-5mwYY7zEgDDDPsanOBTgGQPQHGgJsZiIgYTFmYzExYzAxYjljYjg3ZGZjZWJiYWZmMGVmMGM0ZmQ'}, {'domain': '.douyin.com', 'expiry': 1673363904, 'httpOnly': True, 'name': 'sid_ucp_sso_v1', 'path': '/', 'secure': True, 'value': '1.0.0-KDJhZGY1NzFjNWRhMzkxY2QxNjUyMDdhMDFjZmM2OGExODkwYmQ1ZGMKHQjjhsukhgIQwM-5mwYY7zEgDDDPsanOBTgGQPQHGgJsZiIgYTFmYzExYzAxYjljYjg3ZGZjZWJiYWZmMGVmMGM0ZmQ'}, {'domain': '.douyin.com', 'expiry': 1702739904, 'httpOnly': False, 'name': 'passport_assist_user', 'path': '/', 'secure': True, 'value': 'CjzXpz6-aFYZyJSshOCwwstxVyrxKDHgx8haS-r3eg-Ix6n_Nu4DwjOxKC259JE3sZG3MPKuVcPSuqcmAmsaSAo89U20CSSBw8yVtkFOY9vzEaHfQFGjIjNqs8wfZli_w8WJsOfMNY31uePDLI-bpa2AsjlfQL0Z1oFnoO7vEIX6oA0Yia_WVCIBA_b7tNg%3D'}, {'domain': '.douyin.com', 'expiry': 1699715904, 'httpOnly': True, 'name': 'odin_tt', 'path': '/', 'secure': False, 'value': '9b557c448186766bb41f1c1599d1273cdc3f4f154b3ac3014a2d218e78e7b8c18fbdf5e2e3d867a4c5aeac33d8a4ff57'}, {'domain': '.douyin.com', 'expiry': 1673363902, 'httpOnly': True, 'name': 'ssid_ucp_v1', 'path': '/', 'sameSite': 'None', 'secure': True, 'value': '1.0.0-KDVlZWI4ZmE5ZWQ1ODc2ZGQ0MjU5ZjMwYWIyY2E0YTg4NTIzY2EwZWEKFwjjhsukhgIQv8-5mwYY7zEgDDgGQPQHGgJscSIgYzQ1YWNiMmM5OWJjMDVmYWU3ZjI0NGM5MDQzNWZkMjU'}, {'domain': '.douyin.com', 'expiry': 1673363902, 'httpOnly': True, 'name': 'sid_ucp_v1', 'path': '/', 'secure': True, 'value': '1.0.0-KDVlZWI4ZmE5ZWQ1ODc2ZGQ0MjU5ZjMwYWIyY2E0YTg4NTIzY2EwZWEKFwjjhsukhgIQv8-5mwYY7zEgDDgGQPQHGgJscSIgYzQ1YWNiMmM5OWJjMDVmYWU3ZjI0NGM5MDQzNWZkMjU'}, {'domain': '.douyin.com', 'expiry': 1673363902, 'httpOnly': True, 'name': 'sessionid_ss', 'path': '/', 'sameSite': 'None', 'secure': True, 'value': 'c45acb2c99bc05fae7f244c90435fd25'}, {'domain': '.douyin.com', 'expiry': 1673363902, 'httpOnly': True, 'name': 'sid_tt', 'path': '/', 'secure': False, 'value': 'c45acb2c99bc05fae7f244c90435fd25'}, {'domain': '.douyin.com', 'expiry': 1673363902, 'httpOnly': True, 'name': 'uid_tt_ss', 'path': '/', 'sameSite': 'None', 'secure': True, 'value': '194554730d04e74d3546f14c5351a181'}, {'domain': 'www.douyin.com', 'expiry': 1673363870, 'httpOnly': False, 'name': 's_v_web_id', 'path': '/', 'secure': False, 'value': 'verify_lacn82cu_5iouQAW4_Xl2W_4vIB_B0kq_XNr2Ve7g1KKd'}, {'domain': '.douyin.com', 'expiry': 1699283903, 'httpOnly': True, 'name': 'sid_guard', 'path': '/', 'secure': False, 'value': 'c45acb2c99bc05fae7f244c90435fd25%7C1668179903%7C5183999%7CTue%2C+10-Jan-2023+15%3A18%3A22+GMT'}, {'domain': '.douyin.com', 'expiry': 1670771903, 'httpOnly': True, 'name': 'passport_auth_status_ss', 'path': '/', 'sameSite': 'None', 'secure': True, 'value': '3229dc28f3c954e4644e82adbfd196ba%2C'}, {'domain': '.douyin.com', 'expiry': 1670771903, 'httpOnly': True, 'name': 'passport_auth_status', 'path': '/', 'secure': False, 'value': '3229dc28f3c954e4644e82adbfd196ba%2C'}, {'domain': '.douyin.com', 'expiry': 1673363904, 'httpOnly': True, 'name': 'toutiao_sso_user', 'path': '/', 'secure': False, 'value': 'a1fc11c01b9cb87dfcebbaff0ef0c4fd'}, {'domain': '.douyin.com', 'expiry': 1678547904, 'httpOnly': True, 'name': 'n_mh', 'path': '/', 'secure': False, 'value': 'Khfb_3d2P2Q6Qa4S5hXCYIxEy_l5cgRDDlFW6o97YLk'}, {'domain': '.douyin.com', 'expiry': 1673363902, 'httpOnly': True, 'name': 'uid_tt', 'path': '/', 'secure': False, 'value': '194554730d04e74d3546f14c5351a181'}, {'domain': 'www.douyin.com', 'expiry': 1668784679, 'httpOnly': False, 'name': 'tt_scid', 'path': '/', 'secure': False, 'value': '823JXVZKVmTS2rEPZWXYYjVBfrk9cT4-zDFlbESxZlbLtLQW9KvqlsnEq4dhMDckbc45'}, {'domain': 'www.douyin.com', 'expiry': 1668784708, 'httpOnly': False, 'name': 'msToken', 'path': '/', 'secure': False, 'value': 'fR6yBWBAFr9zbE9PeZ2o39hE4_r8EbTYK9TiACa5RDTsq-gBy6NcHe_eHqo_0xqpIzN9iwAVEYjfm5KlT3T9pUFVCEPIy8kyPUOif3KJyVnJC8bXtENnb0yrnMswlFw='}, {'domain': '.douyin.com', 'expiry': 1673363870, 'httpOnly': False, 'name': 'passport_csrf_token', 'path': '/', 'sameSite': 'None', 'secure': True, 'value': 'acce35bb02060e970db08ec6ea373887'}, {'domain': '.douyin.com', 'expiry': 1668784708, 'httpOnly': False, 'name': 'home_can_add_dy_2_desktop', 'path': '/', 'secure': False, 'value': '%221%22'}, {'domain': '.douyin.com', 'expiry': 1668784669, 'httpOnly': False, 'name': 'strategyABtestKey', 'path': '/', 'secure': False, 'value': '1668179869.729'}, {'domain': '.douyin.com', 'expiry': 1673363904, 'httpOnly': True, 'name': 'toutiao_sso_user_ss', 'path': '/', 'sameSite': 'None', 'secure': True, 'value': 'a1fc11c01b9cb87dfcebbaff0ef0c4fd'}, {'domain': '.douyin.com', 'expiry': 1673363904, 'httpOnly': True, 'name': 'sso_uid_tt_ss', 'path': '/', 'sameSite': 'None', 'secure': True, 'value': 'a938f4e4a1e66029a23196e92300a3cb'}, {'domain': 'www.douyin.com', 'expiry': 1699715868, 'httpOnly': False, 'name': '__ac_signature', 'path': '/', 'secure': False, 'value': '_02B4Z6wo00f01irwsewAAIDCqvJLrkIMgPoq0JVAAOng6c'}, {'domain': '.douyin.com', 'expiry': 1673363870, 'httpOnly': False, 'name': 'passport_csrf_token_default', 'path': '/', 'secure': False, 'value': 'acce35bb02060e970db08ec6ea373887'}, {'domain': '.douyin.com', 'expiry': 1673363904, 'httpOnly': True, 'name': 'sso_uid_tt', 'path': '/', 'secure': False, 'value': 'a938f4e4a1e66029a23196e92300a3cb'}, {'domain': 'www.douyin.com', 'expiry': 1668181667, 'httpOnly': False, 'name': '__ac_nonce', 'path': '/', 'secure': False, 'value': '0636e679b00b7f968e960'}, {'domain': '.douyin.com', 'expiry': 1673363902, 'httpOnly': True, 'name': 'sessionid', 'path': '/', 'secure': False, 'value': 'c45acb2c99bc05fae7f244c90435fd25'}, {'domain': '.douyin.com', 'expiry': 1699283904, 'httpOnly': True, 'name': 'ttwid', 'path': '/', 'secure': False, 'value': '1%7CfZTyB9HMA_PKuT8fr61f19w_t1HHo7-g5sC0J1t86mQ%7C1668179868%7C4f87348cda932d621c3e5adbede771ed8bcbd55d1111d471ca5079e64ed2a44d'}, {'domain': 'www.douyin.com', 'expiry': 1668784678, 'httpOnly': False, 'name': 'ttcid', 'path': '/', 'secure': False, 'value': 'a297f6b1bc1c4b01a9182212017b08f214'}, {'domain': 'www.douyin.com', 'httpOnly': False, 'name': '', 'path': '/video', 'secure': False, 'value': 'douyin.com'}]
    douyin_comment_clawer(douyinid=7073782200499571999,
                          douyincookie=douyincookie,
                          likenumber=50,
                          comment_area_class_name='sX7gMtFl.comment-mainContent',
                          like_class_value='eJuDTubq',
                          block_class_name='RHiEl2d8',
                          reply_comment_class_name='nNNp3deF',
                          name_class_name='Nu66P_ba.NCRZnxVF',
                          comment_first='YzbzCgxU',
                          comment_second='Nu66P_ba.undefined',
                          at_someone_class='B3AsdZT9.DxCF1YBq.VQCdkeKT',
                          time_class_name='L4ozKLf7')
